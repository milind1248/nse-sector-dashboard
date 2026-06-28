"""One-click Excel export — all dashboard data in one workbook."""
import io
import sqlite3
from datetime import date
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

_DB = Path(__file__).resolve().parent.parent / "data" / "nse_dashboard.db"

GREEN_FILL  = PatternFill("solid", fgColor="1B5E20")
RED_FILL    = PatternFill("solid", fgColor="B71C1C")
HEADER_FILL = PatternFill("solid", fgColor="1A237E")


def _apply_pct_colors(ws, col_letters: list[str], start_row: int = 2):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            if cell.column_letter in col_letters:
                try:
                    val = float(str(cell.value or 0).replace("%", "").replace("+", ""))
                    if val > 0:
                        cell.fill = GREEN_FILL
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif val < 0:
                        cell.fill = RED_FILL
                        cell.font = Font(color="FFFFFF", bold=True)
                except Exception:
                    pass


def _write_sheet(wb: openpyxl.Workbook, name: str, df: pd.DataFrame,
                 pct_cols: list[str] = None):
    ws = wb.create_sheet(name)
    for col_num, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_data in dataframe_to_rows(df, index=False, header=False):
        ws.append(row_data)
    if pct_cols:
        col_map = {col: openpyxl.utils.get_column_letter(i + 1)
                   for i, col in enumerate(df.columns)}
        letters = [col_map[c] for c in pct_cols if c in col_map]
        _apply_pct_colors(ws, letters)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)
    return ws


def _db(query: str, params=()) -> pd.DataFrame:
    try:
        con = sqlite3.connect(str(_DB))
        df = pd.read_sql_query(query, con, params=params)
        con.close()
        return df
    except Exception:
        return pd.DataFrame()


def generate_excel_report() -> bytes:
    """Fetch all data from DB + live sources and return Excel bytes."""
    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent))

    from backend.data_ingestion.yfinance_fetcher import (
        fetch_all_sector_prices, compute_pct_returns, _get_close,
    )
    from backend.data_ingestion.nse_fetcher import fetch_fii_dii
    from backend.calculations.indicators import compute_all_indicators
    from backend.calculations.sector_score import compute_sector_score, score_label

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 1. Summary ────────────────────────────────────────────────────────────
    fii_df = fetch_fii_dii(days=120)
    fii_5d = round(fii_df["fii_net"].tail(5).sum(), 2) if not fii_df.empty and "fii_net" in fii_df.columns else "N/A"
    dii_5d = round(fii_df["dii_net"].tail(5).sum(), 2) if not fii_df.empty and "dii_net" in fii_df.columns else "N/A"

    sector_snap = _db("SELECT * FROM daily_sector_snapshot ORDER BY momentum_score DESC")
    strong = int((sector_snap["momentum_score"] >= 70).sum()) if not sector_snap.empty and "momentum_score" in sector_snap.columns else "N/A"
    weak   = int((sector_snap["momentum_score"] <= 30).sum()) if not sector_snap.empty and "momentum_score" in sector_snap.columns else "N/A"

    summary_rows = [
        ["Generated On",                date.today().isoformat()],
        ["Total Sectors tracked",        len(sector_snap) if not sector_snap.empty else "N/A"],
        ["Strong Sectors (Score ≥ 70)",  strong],
        ["Weak Sectors (Score ≤ 30)",    weak],
        ["FII Net Last 5 Days (₹ Cr)",   fii_5d],
        ["DII Net Last 5 Days (₹ Cr)",   dii_5d],
        ["FII/DII History Days",         len(fii_df)],
        ["Index Stocks tracked",         len(_db("SELECT * FROM sector_intelligence"))],
    ]
    _write_sheet(wb, "1_Summary",
                 pd.DataFrame(summary_rows, columns=["Metric", "Value"]))

    # ── 2. Sector Analysis (daily snapshot) ───────────────────────────────────
    if not sector_snap.empty:
        pct_cols = [c for c in sector_snap.columns
                    if any(x in c for x in ("pct_", "rs_", "ad_ratio"))]
        _write_sheet(wb, "2_Sector_Analysis", sector_snap, pct_cols=pct_cols)

    # ── 3. Sector Heatmap (live % returns) ───────────────────────────────────
    sector_prices = fetch_all_sector_prices()
    hm_rows = []
    sector_live_rows = []
    for s, df in sector_prices.items():
        if df is None or df.empty:
            continue
        rets  = compute_pct_returns(df)
        indic = compute_all_indicators(df)
        close_s = _get_close(df)
        close   = float(close_s.iloc[-1]) if close_s is not None and not close_s.empty else None
        score   = compute_sector_score(
            rs_vs_nifty=None, pct_1w=rets.get("pct_1w"), pct_1m=rets.get("pct_1m"),
            rsi_14=indic.get("rsi_14"), close=close, ema_200=indic.get("ema_200"),
        )
        lbl, _ = score_label(score)
        hm_rows.append({"Sector": s,
                         "1W%": rets.get("pct_1w"), "2W%": rets.get("pct_2w"),
                         "1M%": rets.get("pct_1m"), "3M%": rets.get("pct_3m"),
                         "6M%": rets.get("pct_6m"), "1Y%": rets.get("pct_1y")})
        sector_live_rows.append({
            "Sector": s, "Score": score, "Signal": lbl,
            "RSI_14":  round(indic.get("rsi_14")  or 0, 2),
            "EMA_20":  round(indic.get("ema_20")  or 0, 2),
            "EMA_50":  round(indic.get("ema_50")  or 0, 2),
            "EMA_200": round(indic.get("ema_200") or 0, 2),
            "1W%": rets.get("pct_1w"),  "2W%": rets.get("pct_2w"),
            "1M%": rets.get("pct_1m"),  "3M%": rets.get("pct_3m"),
            "6M%": rets.get("pct_6m"),  "1Y%": rets.get("pct_1y"),
        })

    if hm_rows:
        hm_df = pd.DataFrame(hm_rows).set_index("Sector").reset_index()
        _write_sheet(wb, "3_Heatmap", hm_df,
                     pct_cols=[c for c in hm_df.columns if "%" in c])

    if sector_live_rows:
        live_df = pd.DataFrame(sector_live_rows).sort_values("Score", ascending=False)
        _write_sheet(wb, "4_Sector_Live_Scores", live_df,
                     pct_cols=[c for c in live_df.columns if "%" in c])

    # ── 4. Index Stocks with Weightage ────────────────────────────────────────
    idx_stocks = _db(
        "SELECT sector, index_name, index_display, company_name, symbol, "
        "       industry, weightage_pct, market_cap_cr, weight_source "
        "FROM sector_intelligence ORDER BY sector, index_name, weightage_pct DESC"
    )
    if not idx_stocks.empty:
        _write_sheet(wb, "5_Index_Stocks", idx_stocks)

    # ── 5. Stock Snapshot (all stocks with indicators) ────────────────────────
    stock_snap = _db(
        "SELECT date, sector, symbol, name, close, market_cap, "
        "       pct_1d, pct_1w, pct_1m, pct_3m, pct_6m, pct_1y, "
        "       rsi_14, ema_20, ema_50, ema_200, macd, "
        "       fii_holding_pct, dii_holding_pct, promoter_pct, mf_pct, "
        "       high_52w, low_52w, rs_vs_nifty, momentum_score "
        "FROM daily_stock_snapshot ORDER BY sector, momentum_score DESC"
    )
    if not stock_snap.empty:
        pct_cols = [c for c in stock_snap.columns
                    if "pct_" in c or "rs_" in c]
        _write_sheet(wb, "6_Stock_Snapshot", stock_snap, pct_cols=pct_cols)

    # ── 6. FII / DII Daily Flow ───────────────────────────────────────────────
    if not fii_df.empty:
        _write_sheet(wb, "7_FII_DII_Daily",
                     fii_df.sort_values("date", ascending=False))

    # ── 7. NSDL Sector FII (fortnightly) ─────────────────────────────────────
    nsdl_df = _db(
        "SELECT report_date, nsdl_sector, sector, "
        "       auc_prev_eq, net_prev_eq, net_curr_eq, auc_curr_eq, "
        "       auc_change, auc_pct_change, net_flow_change, signal "
        "FROM nsdl_fii_sector ORDER BY report_date DESC, net_curr_eq DESC"
    )
    if not nsdl_df.empty:
        _write_sheet(wb, "8_NSDL_Sector_FII", nsdl_df,
                     pct_cols=["auc_pct_change"])

    # ── 8. Market Breadth ─────────────────────────────────────────────────────
    breadth_df = _db("SELECT * FROM market_breadth_daily ORDER BY date DESC")
    if not breadth_df.empty:
        _write_sheet(wb, "9_Market_Breadth", breadth_df)

    # ── 9. Smart Money Screener (latest Buying signals) ───────────────────────
    sm_screener = _db("""
        WITH latest AS (
            SELECT symbol, MAX(trade_date) AS last_date
            FROM smart_money_history
            WHERE close_price IS NOT NULL
            GROUP BY symbol
        ),
        avg90 AS (
            SELECT symbol,
                   AVG(dlv_pct) AS avg_dlv,
                   AVG(action)  AS avg_action
            FROM smart_money_history
            WHERE close_price IS NOT NULL
            GROUP BY symbol
        ),
        last_row AS (
            SELECT h.symbol, h.trade_date, h.close_price, h.pct_price_chg,
                   h.dlv_pct, h.action, h.futures_oi, h.oi_change, h.pct_oi_chg
            FROM smart_money_history h
            JOIN latest l ON h.symbol = l.symbol AND h.trade_date = l.last_date
        )
        SELECT r.symbol, r.trade_date AS signal_date,
               r.close_price, r.pct_price_chg AS pct_price_chg,
               ROUND(r.dlv_pct, 2) AS dlv_pct,
               ROUND(a.avg_dlv, 2) AS avg_dlv_90d,
               ROUND(r.action, 2)  AS action,
               ROUND(a.avg_action, 2) AS avg_action_90d,
               r.futures_oi, r.oi_change, r.pct_oi_chg
        FROM last_row r
        JOIN avg90 a ON r.symbol = a.symbol
        WHERE r.dlv_pct > a.avg_dlv AND r.action > a.avg_action
        ORDER BY r.dlv_pct DESC
    """)
    if not sm_screener.empty:
        sm_screener.insert(0, "Smart_Money_Signal", "Buying")
        _write_sheet(wb, "10_Smart_Money_Screener", sm_screener,
                     pct_cols=["pct_price_chg", "dlv_pct", "avg_dlv_90d", "pct_oi_chg"])

    # ── 10. Smart Money History (last 90 days, all FNO stocks) ────────────────
    sm_history = _db("""
        SELECT symbol, trade_date, close_price, pct_price_chg,
               dlv_pct, action, futures_oi, oi_change, pct_oi_chg,
               CASE WHEN dlv_pct > AVG(dlv_pct) OVER (PARTITION BY symbol)
                         AND action > AVG(action) OVER (PARTITION BY symbol)
                    THEN 'Buying' ELSE '' END AS smart_money_signal
        FROM smart_money_history
        WHERE close_price IS NOT NULL
        ORDER BY symbol, trade_date DESC
    """)
    if not sm_history.empty:
        _write_sheet(wb, "11_Smart_Money_History", sm_history,
                     pct_cols=["pct_price_chg", "dlv_pct", "pct_oi_chg"])

    # ── 11. AI Forecast Signals (Prophet + XGBoost for all stocks) ───────────
    ai_df = _db("""
        SELECT symbol            AS "Symbol",
               sector            AS "Sector",
               ROUND(price, 2)   AS "Price (Rs)",
               xgb_direction     AS "XGB Direction",
               ROUND(xgb_prob * 100, 1) AS "XGB Probability %",
               xgb_signal        AS "Signal",
               ROUND(xgb_accuracy, 1)   AS "Backtest Accuracy %",
               prophet_trend     AS "Prophet Trend",
               ROUND(prophet_trend_pct, 2) AS "Prophet % Change (30d)",
               arima_direction   AS "ARIMA Trend",
               ROUND(arima_trend_pct, 2)   AS "ARIMA % Change (30d)",
               scan_date         AS "Scan Date"
        FROM ai_forecast_cache
        ORDER BY xgb_prob DESC
    """)
    if not ai_df.empty:
        _write_sheet(wb, "12_AI_Forecast_Signals", ai_df,
                     pct_cols=["XGB Probability %", "Backtest Accuracy %", "Prophet % Change (30d)", "ARIMA % Change (30d)"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
